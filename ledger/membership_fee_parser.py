import re
import argparse
import pandas as pd
import openpyxl
from openpyxl.styles import Font

TEMPLATE_PATH = 'templates/ledger_format.xlsx'
LEDGER_PATH = 'ledger.xlsx'
SOURCE_HEADER_ROW = 1   # 0-indexed: row 2 in Excel
SOURCE_COLS = 'C:I'     # 월, 날짜, 내용, 이름, 비고, 입/출, 잔액


def _ym_from_date(date_str):
    """날짜 문자열에서 YYYYMM 정수 추출: '2025.11.01 ...' -> 202511"""
    m = re.match(r'(\d{4})\.(\d{2})\.', str(date_str))
    return int(m.group(1)) * 100 + int(m.group(2)) if m else None


def _parse_period(ym_str):
    """'YYYY-MM' 문자열을 YYYYMM 정수로 변환"""
    m = re.match(r'^(\d{4})-(\d{1,2})$', ym_str.strip())
    if not m:
        raise ValueError(f"기간 형식이 올바르지 않습니다: '{ym_str}' (예: 2025-11)")
    return int(m.group(1)) * 100 + int(m.group(2))


def parse_source(file_path):
    all_sheets = pd.read_excel(
        file_path, sheet_name=None, header=SOURCE_HEADER_ROW, usecols=SOURCE_COLS
    )
    year_pattern = re.compile(r'^\d{4}년$')
    frames = [df for name, df in all_sheets.items() if year_pattern.match(name)]
    if not frames:
        raise ValueError("연도별 시트(YYYY년)를 찾을 수 없습니다.")
    df = pd.concat(frames, ignore_index=True)
    df = df.dropna(subset=['날짜']).reset_index(drop=True)
    return df


def filter_by_period(df, start_ym, end_ym):
    """YYYYMM 범위로 필터링 후 날짜순 정렬"""
    yms = df['날짜'].apply(_ym_from_date)
    filtered = df[(yms >= start_ym) & (yms <= end_ym)].copy()
    filtered = filtered.sort_values('날짜').reset_index(drop=True)
    return filtered


def _clear_data_rows(ws, data_start_row):
    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def write_to_ledger(df, template_path, output_path, start_ym, end_ym):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    start_balance = df['잔액'].iloc[0] - df['입/출'].iloc[0]
    end_balance = df['잔액'].iloc[-1]

    # 잔액 기준일 텍스트 업데이트
    ws['E4'] = f"{start_ym // 100}년 {start_ym % 100}월 1일자 기준 동아리 통장 잔액"
    ws['F4'] = start_balance
    ws['E5'] = f"{end_ym // 100}년 {end_ym % 100}월 말일자 기준 동아리 통장 잔액"
    ws['F5'] = end_balance

    # 기존 데이터 행 초기화 후 덮어쓰기
    DATA_START_ROW = 8
    _clear_data_rows(ws, DATA_START_ROW)

    for i, row in df.iterrows():
        excel_row = DATA_START_ROW + i
        amount = row['입/출']
        is_income = amount > 0
        종류 = '입금' if is_income else '출금'
        font = Font(color='000000') if is_income else Font(color='FF0000')

        cells = [
            ws.cell(row=excel_row, column=2, value=i + 1),
            ws.cell(row=excel_row, column=3, value=str(row['날짜'])),
            ws.cell(row=excel_row, column=4, value=종류),
            ws.cell(row=excel_row, column=5, value=str(row['내용'])),
            ws.cell(row=excel_row, column=6, value=abs(amount)),
            ws.cell(row=excel_row, column=7, value=row['잔액']),
        ]
        for cell in cells:
            cell.font = font

    wb.save(output_path)
    print(f"저장 완료: {output_path} ({len(df)}개 행)")


def update_ledger_sheet1(df, ledger_path):
    """ledger.xlsx의 Sheet1을 필터링된 데이터로 덮어씀.
    Sheet1 형식: 날짜 | 종류(내용) | 금액(부호 포함) | 잔액
    """
    wb = openpyxl.load_workbook(ledger_path)
    ws = wb['Sheet1']

    # 헤더(1행) 제외하고 기존 데이터 초기화
    _clear_data_rows(ws, 2)

    for i, row in df.iterrows():
        excel_row = 2 + i
        ws.cell(row=excel_row, column=1, value=str(row['날짜']))
        ws.cell(row=excel_row, column=2, value=str(row['내용']))
        ws.cell(row=excel_row, column=3, value=row['입/출'])
        ws.cell(row=excel_row, column=4, value=row['잔액'])

    wb.save(ledger_path)
    print(f"ledger.xlsx Sheet1 업데이트 완료 ({len(df)}개 행)")


def run(source_path, start, end, output_path=None, template_path=None, ledger_path=None):
    start_ym = _parse_period(start)
    end_ym = _parse_period(end)

    if start_ym > end_ym:
        raise ValueError(f"시작 기간({start})이 종료 기간({end})보다 늦습니다.")

    if output_path is None:
        output_path = f'ledger_{start.replace("-", "")}_{end.replace("-", "")}.xlsx'

    df = parse_source(source_path)
    df_filtered = filter_by_period(df, start_ym, end_ym)

    if df_filtered.empty:
        print(f"경고: {start}~{end} 범위의 데이터가 없습니다.")
        return

    write_to_ledger(df_filtered, template_path or TEMPLATE_PATH, output_path, start_ym, end_ym)
    update_ledger_sheet1(df_filtered, ledger_path or LEDGER_PATH)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='재학생 회비 관리 문서 → 장부 변환')
    parser.add_argument('source', help='소스 파일 경로 (재학생 회비 관리 문서_날짜.xlsx)')
    parser.add_argument('start', help='시작 기간 (예: 2025-11)')
    parser.add_argument('end', help='종료 기간 (예: 2026-02)')
    parser.add_argument('--output', help='출력 파일 경로 (기본: ledger_YYYYMM_YYYYMM.xlsx)')
    args = parser.parse_args()

    run(args.source, args.start, args.end, args.output)

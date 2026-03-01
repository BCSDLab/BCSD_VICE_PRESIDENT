#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
신한 거래내역 → 재학생 회비 관리 문서 자동 기입

자동 기입: 날짜, 이름(거래내역 내용), 입/출, 잔액, 소계/합계 수식
수동 기입: 내용(E열), 비고(G열)

사용법:
    python fill_ledger.py                        # 가장 최신 거래내역 파일 사용
    python fill_ledger.py 신한_거래내역/신한_거래내역_2602.xlsx
    python fill_ledger.py --force                # 이미 기입된 월도 덮어쓰기
"""

import io
import os
import re
import sys
import argparse
from collections import Counter
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter
from common.google_drive import (
    _extract_drive_folder_id,
    _extract_sheet_id,
    _download_request_to_tempfile,
)

try:
    from googleapiclient.errors import HttpError as _HttpError
except ImportError:
    _HttpError = OSError  # type: ignore[assignment,misc]

# ============================================================================
# Constants
# ============================================================================

# 관리 문서 컬럼 (1-based)
COL_MONTH = 3   # C: 월
COL_DATE = 4    # D: 날짜
COL_DESC = 5    # E: 내용 (수동 기입)
COL_NAME = 6    # F: 이름
COL_NOTE = 7    # G: 비고 (수동 기입)
COL_AMOUNT = 8  # H: 입/출
COL_BALANCE = 9 # I: 잔액

# 수식 작성에 사용할 열 문자 (COL_* 상수에서 파생)
_L_DESC    = get_column_letter(COL_DESC)    # E
_L_NOTE    = get_column_letter(COL_NOTE)    # G
_L_AMOUNT  = get_column_letter(COL_AMOUNT)  # H


# ============================================================================
# Google Auth
# ============================================================================

GOOGLE_TOKEN_FILE = '.google_token.json'
_GOOGLE_SCOPES = ['https://www.googleapis.com/auth/drive']


def _get_credentials():
    """OAuth 인증 credentials 반환. 토큰은 GOOGLE_TOKEN_FILE에 캐시."""
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
        from google_auth_oauthlib.flow import InstalledAppFlow
    except ImportError as err:
        raise ImportError(
            "Google 연동에 필요한 패키지가 없습니다: pip install google-auth google-auth-oauthlib google-api-python-client"
        ) from err

    creds = None
    if os.path.exists(GOOGLE_TOKEN_FILE):
        try:
            creds = Credentials.from_authorized_user_file(GOOGLE_TOKEN_FILE, _GOOGLE_SCOPES)
        except (ValueError, OSError) as e:
            print(f"[WARNING] 토큰 파일이 손상되어 재인증합니다: {e}")
            creds = None

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception as e:
                print(f"[WARNING] 토큰 갱신 실패, 재인증합니다: {e}")
                creds = None
        if not creds or not creds.valid:
            secret_json = os.getenv('GOOGLE_OAUTH_CLIENT_JSON') or os.getenv('GOOGLE_SECRET_JSON')
            if not secret_json:
                raise ValueError("[ERROR] GOOGLE_OAUTH_CLIENT_JSON 환경변수가 설정되지 않았습니다.")
            flow = InstalledAppFlow.from_client_secrets_file(secret_json, _GOOGLE_SCOPES)
            creds = flow.run_local_server(port=0)
            if not creds or not creds.valid:
                raise RuntimeError("OAuth 인증에 실패했습니다.")
        fd = os.open(GOOGLE_TOKEN_FILE, os.O_WRONLY | os.O_CREAT | os.O_TRUNC, 0o600)
        with os.fdopen(fd, 'w') as f:
            f.write(creds.to_json())

    return creds


def _get_drive_service():
    """Drive v3 서비스 객체 반환."""
    from googleapiclient.discovery import build
    return build('drive', 'v3', credentials=_get_credentials())


def _get_sheets_service():
    """Sheets v4 서비스 객체 반환."""
    from googleapiclient.discovery import build
    return build('sheets', 'v4', credentials=_get_credentials())


# ============================================================================
# Google Drive integration
# ============================================================================

def download_sheet_as_xlsx(url):
    """
    Google Sheets 문서를 xlsx 파일로 내보내 임시 파일에 저장.

    Returns:
        (spreadsheet_id, tmp_path) — 호출자가 tmp_path 삭제 책임
    """
    spreadsheet_id = _extract_sheet_id(url)
    if not spreadsheet_id:
        raise ValueError(f"Google Sheets URL에서 ID를 파싱할 수 없습니다: {url}")
    drive = _get_drive_service()
    request = drive.files().export_media(
        fileId=spreadsheet_id,
        mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    return spreadsheet_id, _download_request_to_tempfile(request, suffix='.xlsx')


def _find_latest_transaction_in_folder(drive, folder_id):
    """폴더 내 신한_거래내역_YYMM.xlsx 파일 중 가장 최신 파일의 (file_id, name) 반환."""
    pattern = re.compile(r'신한_거래내역_(\d{4})\.xlsx$')
    files = []
    page_token = None
    while True:
        kwargs = dict(
            q=f"'{folder_id}' in parents and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' and trashed=false",
            fields='nextPageToken, files(id, name)',
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        )
        if page_token:
            kwargs['pageToken'] = page_token
        result = drive.files().list(**kwargs).execute()
        for f in result.get('files', []):
            m = pattern.search(f['name'])
            if m:
                s = m.group(1)
                if 1 <= int(s[2:]) <= 12:
                    files.append((int(s), f['name'], f['id']))
        page_token = result.get('nextPageToken')
        if not page_token:
            break
    if not files:
        raise FileNotFoundError("폴더에서 신한_거래내역_YYMM.xlsx 파일을 찾을 수 없습니다.")
    files.sort(key=lambda x: x[0])
    _, name, file_id = files[-1]
    return file_id, name


def download_transaction_from_drive(url):
    """
    Google Drive 폴더에서 최신 거래내역 xlsx 파일 다운로드.

    Returns:
        (original_filename, tmp_path) — 호출자가 tmp_path를 사용 후 삭제 책임
    """
    folder_id = _extract_drive_folder_id(url)
    if not folder_id:
        raise ValueError(f"Google Drive 폴더 URL에서 ID를 파싱할 수 없습니다: {url}")

    drive = _get_drive_service()
    file_id, original_name = _find_latest_transaction_in_folder(drive, folder_id)

    request = drive.files().get_media(fileId=file_id, supportsAllDrives=True)
    return original_name, _download_request_to_tempfile(request, suffix='.xlsx')


# ============================================================================
# Parsing
# ============================================================================

def parse_transaction_file(filepath):
    """
    거래내역 엑셀 파일 파싱.

    Returns:
        list of (date_str, amount, name, balance)
        - amount: 입금이면 양수, 출금이면 음수
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return []
        transactions = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 6:
                continue
            no, date_str, deposit, withdrawal, name, balance = row[:6]
            if no is None:
                continue
            def _to_number(value):
                if isinstance(value, (int, float)):
                    return float(value)
                if isinstance(value, str):
                    raw = value.replace(',', '').strip()
                    try:
                        return float(raw) if raw else 0.0
                    except ValueError:
                        return 0.0
                return 0.0

            deposit = _to_number(deposit)
            withdrawal = _to_number(withdrawal)
            if deposit > 0:
                amount = deposit
            elif withdrawal > 0:
                amount = -withdrawal
            else:
                continue
            if isinstance(date_str, datetime):
                date_value = date_str.strftime('%Y.%m.%d %H:%M:%S')
            else:
                date_value = str(date_str)
            if isinstance(balance, (int, float)):
                safe_balance = balance
            elif isinstance(balance, str):
                raw = balance.replace(',', '').strip()
                try:
                    safe_balance = float(raw) if raw else None
                except ValueError:
                    safe_balance = None
            else:
                safe_balance = None
            transactions.append((date_value, amount, str(name) if name else "", safe_balance))
    finally:
        wb.close()

    return transactions


def get_year_month_from_filename(filepath):
    """
    파일명에서 연도/월 추출.

    Examples:
        신한_거래내역_2601.xlsx → (2026, 1)
        신한_거래내역_2512.xlsx → (2025, 12)
    """
    basename = os.path.basename(filepath)
    match = re.search(r'_(\d{2})(\d{2})\.xlsx$', basename)
    if not match:
        raise ValueError(f"파일명에서 연도/월을 파싱할 수 없습니다: {basename}")
    yy, mm = int(match.group(1)), int(match.group(2))
    if not 1 <= mm <= 12:
        raise ValueError(f"파일명의 월 값이 유효하지 않습니다 (mm={mm}): {basename}")
    return 2000 + yy, mm


# ============================================================================
# Sheets API helpers
# ============================================================================

def _get_sheet_gid(sheets, spreadsheet_id, sheet_name):
    """시트 이름으로 sheetId(gid) 반환."""
    result = sheets.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields='sheets.properties',
    ).execute()
    for sheet in result.get('sheets', []):
        props = sheet.get('properties', {})
        if props.get('title') == sheet_name:
            return props['sheetId']
    raise ValueError(f"시트 '{sheet_name}'를 찾을 수 없습니다.")


def _read_col_c(sheets, spreadsheet_id, sheet_name):
    """C열 전체 값 읽기. 리스트[0] = 행1, 빈 셀은 '' 반환."""
    result = sheets.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=f"'{sheet_name}'!C:C",
    ).execute()
    rows = result.get('values', [])
    return [row[0] if row else '' for row in rows]


def _find_month_section_api(sheets, spreadsheet_id, sheet_name, month):
    """
    월 섹션의 header_row, sogyeyu_row 반환 (1-based).

    Returns:
        (header_row, sogyeyu_row) or (None, None)
    """
    month_label = f"{month}월"
    col_c = _read_col_c(sheets, spreadsheet_id, sheet_name)

    header_row = None
    for i, val in enumerate(col_c):
        if val == month_label:
            header_row = i + 1  # 1-based
            break

    if header_row is None:
        return None, None

    sogyeyu_row = None
    for i in range(header_row, len(col_c)):  # header_row 인덱스(0-based) = header_row+1 행(1-based) 부터
        if col_c[i] == '소계':
            sogyeyu_row = i + 1  # 1-based
            break

    return header_row, sogyeyu_row


def _is_month_filled_api(sheets, spreadsheet_id, sheet_name, header_row):
    """월 섹션이 이미 기입되었는지 확인 (날짜 셀 기준)."""
    col_d = get_column_letter(COL_DATE)
    result = sheets.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=f"'{sheet_name}'!{col_d}{header_row}",
    ).execute()
    return bool(result.get('values'))


def _read_jan_template(sheets, spreadsheet_id, sheet_name):
    """
    1월 섹션 범위와 날짜 저장 방식 반환.

    Returns:
        (jan_header, jan_sogyeyu, date_is_serial)
        - jan_header, jan_sogyeyu: 1-based 행 번호 (None이면 1월 데이터 없음)
        - date_is_serial: True이면 날짜가 date serial(숫자), False이면 문자열
    """
    jan_header, jan_sogyeyu = _find_month_section_api(sheets, spreadsheet_id, sheet_name, 1)
    if jan_header is None or jan_sogyeyu is None:
        return None, None, False
    if not _is_month_filled_api(sheets, spreadsheet_id, sheet_name, jan_header):
        return None, None, False

    # D열(날짜) 셀의 userEnteredValue로 저장 방식 판별 (중간 행 기준)
    sample_row = jan_header + 1 if jan_sogyeyu - jan_header > 1 else jan_header
    col_d = get_column_letter(COL_DATE)
    result = sheets.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        ranges=[f"'{sheet_name}'!{col_d}{sample_row}"],
        fields='sheets.data.rowData.values.userEnteredValue',
        includeGridData=True,
    ).execute()
    try:
        uev = (result['sheets'][0]['data'][0]['rowData'][0]
               ['values'][0].get('userEnteredValue', {}))
        date_is_serial = 'numberValue' in uev
    except (KeyError, IndexError):
        date_is_serial = False

    return jan_header, jan_sogyeyu, date_is_serial


def _date_str_to_sheets_serial(date_str):
    """'YYYY.MM.DD[...]' 문자열을 Google Sheets date serial(숫자)로 변환."""
    from datetime import date
    date_only = date_str[:10]  # 시간 부분 무시
    y, m, d = (int(x) for x in date_only.split('.'))
    return (date(y, m, d) - date(1970, 1, 1)).days + 25569


def _check_manual_entries(sheets, spreadsheet_id, sheet_name, start_row, end_row):
    """삭제될 행에 수동 기입 항목(E, G열)이 있으면 경고."""
    if start_row > end_row:
        return
    for col, col_name in [(COL_DESC, 'E열'), (COL_NOTE, 'G열')]:
        col_letter = get_column_letter(col)
        result = sheets.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_name}'!{col_letter}{start_row}:{col_letter}{end_row}",
        ).execute()
        for i, row_vals in enumerate(result.get('values', [])):
            if row_vals and row_vals[0]:
                print(f"[WARNING] 행 {start_row + i}의 수동 기입 항목({col_name})이 삭제됩니다.")


# ============================================================================
# Sheet fill (Sheets API)
# ============================================================================

def fill_month_api(sheets, spreadsheet_id, sheet_name, gid, month, transactions,
                   force=False, receipt_map=None):
    """
    Sheets API를 사용하여 월 거래내역 기입.

    Returns:
        True  — 기입 완료
        None  — 이미 기입된 월을 건너뜀 (성공적 no-op)
        False — 오류
    """
    month_label = f"{month}월"
    header_row, sogyeyu_row = _find_month_section_api(sheets, spreadsheet_id, sheet_name, month)

    if header_row is None or sogyeyu_row is None:
        print(f"[ERROR] {month_label} 섹션을 찾을 수 없습니다.")
        return False

    if _is_month_filled_api(sheets, spreadsheet_id, sheet_name, header_row):
        if not force:
            print(f"[WARNING] {month_label} 데이터가 이미 존재합니다. 건너뜁니다. (덮어쓰려면 --force 사용)")
            return None
        print(f"[INFO] {month_label} 기존 데이터를 덮어씁니다.")

    tx_count = len(transactions)
    if tx_count == 0:
        print(f"[WARNING] {month_label} 거래 내역이 없습니다.")
        return True

    placeholder_rows = sogyeyu_row - header_row
    delta = tx_count - placeholder_rows
    new_sogyeyu_row = sogyeyu_row + delta
    data_end = header_row + tx_count - 1

    # 삭제될 행에 수동 기입 항목 경고
    if delta < 0:
        _check_manual_entries(
            sheets, spreadsheet_id, sheet_name,
            header_row + tx_count, sogyeyu_row - 1,
        )

    # ── 구조 변경 ──────────────────────────────────────────────────────────
    struct_requests = []

    # 1. C열 병합 해제
    struct_requests.append({
        'unmergeCells': {
            'range': {
                'sheetId': gid,
                'startRowIndex': header_row - 1,
                'endRowIndex': sogyeyu_row - 1,
                'startColumnIndex': COL_MONTH - 1,
                'endColumnIndex': COL_MONTH,
            },
        },
    })

    # 2. 행 삽입/삭제
    #    삽입 시 inheritFromBefore=True: 바로 위 기존 행의 서식을 그대로 복사.
    #    별도 서식 복사 없이 기존 플레이스홀더 행 서식이 자동 적용됨.
    if delta > 0:
        struct_requests.append({
            'insertDimension': {
                'range': {
                    'sheetId': gid,
                    'dimension': 'ROWS',
                    'startIndex': sogyeyu_row - 1,
                    'endIndex': sogyeyu_row - 1 + delta,
                },
                'inheritFromBefore': True,
            },
        })
    elif delta < 0:
        struct_requests.append({
            'deleteDimension': {
                'range': {
                    'sheetId': gid,
                    'dimension': 'ROWS',
                    'startIndex': header_row + tx_count - 1,
                    'endIndex': header_row + placeholder_rows - 1,
                },
            },
        })

    # 3. 데이터 행 D~I 테두리 명시적 복원 (이전 실행에서 손상된 서식 복구)
    _data_range = {
        'sheetId': gid,
        'startRowIndex': header_row - 1,
        'endRowIndex': data_end,
    }
    _solid   = {'style': 'SOLID'}
    _solid_m = {'style': 'SOLID_MEDIUM'}
    struct_requests.append({
        'updateBorders': {
            'range': {**_data_range,
                      'startColumnIndex': COL_DATE - 1,
                      'endColumnIndex': COL_BALANCE},
            'top': _solid_m,          # 섹션 첫 행 상단
            'bottom': _solid,
            'left': _solid,
            'right': _solid,
            'innerHorizontal': _solid,
            'innerVertical': _solid,
        },
    })
    # I열 우측 테두리만 SOLID_MEDIUM으로 덮어쓰기
    struct_requests.append({
        'updateBorders': {
            'range': {**_data_range,
                      'startColumnIndex': COL_BALANCE - 1,
                      'endColumnIndex': COL_BALANCE},
            'right': _solid_m,
        },
    })
    # E열 textFormat 삭제: 명시적 폰트 색을 완전히 제거해 =HYPERLINK() 기본 파란색으로 표시.
    # cell에 textFormat을 포함하지 않으면 fields 마스크가 해당 필드를 unset(삭제)으로 처리.
    struct_requests.append({
        'repeatCell': {
            'range': {**_data_range,
                      'startColumnIndex': COL_DESC - 1,
                      'endColumnIndex': COL_DESC},
            'cell': {'userEnteredFormat': {}},
            'fields': 'userEnteredFormat.textFormat',
        },
    })

    # 4. C열 재병합
    struct_requests.append({
        'mergeCells': {
            'range': {
                'sheetId': gid,
                'startRowIndex': header_row - 1,
                'endRowIndex': new_sogyeyu_row - 1,
                'startColumnIndex': COL_MONTH - 1,
                'endColumnIndex': COL_MONTH,
            },
            'mergeType': 'MERGE_ALL',
        },
    })

    sheets.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={'requests': struct_requests},
    ).execute()

    # ── 값 기입 ────────────────────────────────────────────────────────────
    col_c = get_column_letter(COL_MONTH)
    col_d = get_column_letter(COL_DATE)
    col_e = get_column_letter(COL_DESC)
    col_f = get_column_letter(COL_NAME)
    col_h = get_column_letter(COL_AMOUNT)
    col_i = get_column_letter(COL_BALANCE)

    raw_data = [
        {'range': f"'{sheet_name}'!{col_c}{header_row}",
         'values': [[month_label]]},
        {'range': f"'{sheet_name}'!{col_d}{header_row}:{col_d}{data_end}",
         'values': [[ds] for ds, *_ in transactions]},
        {'range': f"'{sheet_name}'!{col_f}{header_row}:{col_f}{data_end}",
         'values': [[name] for _, _, name, _ in transactions]},
        {'range': f"'{sheet_name}'!{col_h}{header_row}:{col_h}{data_end}",
         'values': [[amount] for _, amount, *_ in transactions]},
        {'range': f"'{sheet_name}'!{col_i}{header_row}:{col_i}{data_end}",
         'values': [[bal if bal is not None else ''] for _, _, _, bal in transactions]},
    ]
    sheets.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={'valueInputOption': 'RAW', 'data': raw_data},
    ).execute()

    # 수식은 USER_ENTERED로 별도 기입
    formula_data = [
        {'range': f"'{sheet_name}'!C{new_sogyeyu_row}:I{new_sogyeyu_row}",
         'values': [[
             '소계', '입금',
             f'=SUMIF({_L_AMOUNT}{header_row}:{_L_AMOUNT}{data_end},">0")',
             '출금',
             f'=SUMIF({_L_AMOUNT}{header_row}:{_L_AMOUNT}{data_end},"<0")*-1',
             '합계',
             f'={_L_DESC}{new_sogyeyu_row}-{_L_NOTE}{new_sogyeyu_row}',
         ]]},
    ]

    # E열: 출금+영수증 매칭 시에만 HYPERLINK 수식 기입 (수동 기입 셀은 건드리지 않음)
    if receipt_map:
        for i, (date_str, amount, *_) in enumerate(transactions):
            if amount < 0:
                key = (date_str[:10], int(abs(amount)))
                if key in receipt_map:
                    title, url = receipt_map[key]
                    safe_url = url.replace('"', '%22')
                    safe_title = title.replace('"', '""')
                    formula_data.append({
                        'range': f"'{sheet_name}'!{col_e}{header_row + i}",
                        'values': [[f'=HYPERLINK("{safe_url}","{safe_title}")']],
                    })

    sheets.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={'valueInputOption': 'USER_ENTERED', 'data': formula_data},
    ).execute()

    print(f"[INFO] {month_label} 거래 {tx_count}건 기입 완료 (행 {header_row}~{data_end}, 소계 행 {new_sogyeyu_row})")
    return True


def update_total_formula_api(sheets, spreadsheet_id, sheet_name):
    """합계 행의 SUM 수식을 현재 소계 행 위치에 맞게 갱신."""
    col_c = _read_col_c(sheets, spreadsheet_id, sheet_name)

    total_row = None
    for i, val in enumerate(col_c):
        if val == '합계':
            total_row = i + 1  # 1-based
            break

    if total_row is None:
        return

    sogyeyu_rows = [
        i + 1
        for i, val in enumerate(col_c[:total_row - 1])
        if val == '소계'
    ]
    if not sogyeyu_rows:
        return

    sum_e = ','.join(f'{_L_DESC}{r}' for r in sogyeyu_rows)
    sum_g = ','.join(f'{_L_NOTE}{r}' for r in sogyeyu_rows)

    sheets.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            'valueInputOption': 'USER_ENTERED',
            'data': [{
                'range': f"'{sheet_name}'!C{total_row}:I{total_row}",
                'values': [[
                    '합계',
                    '입금',
                    f'=SUM({sum_e})',
                    '출금',
                    f'=SUM({sum_g})',
                    '합계',
                    f'={_L_DESC}{total_row}-{_L_NOTE}{total_row}',
                ]],
            }],
        },
    ).execute()

    print(f"[INFO] 합계 행({total_row}) 수식 갱신 완료")


# ============================================================================
# Receipt matching (Google Drive)
# ============================================================================

def _find_drive_subfolder(drive, parent_id, name):
    """Drive 폴더 내 이름이 일치하는 하위 폴더 ID 반환. 없으면 None."""
    result = drive.files().list(
        q=(
            f"'{parent_id}' in parents"
            f" and name='{name}'"
            " and mimeType='application/vnd.google-apps.folder'"
            " and trashed=false"
        ),
        fields='files(id)',
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()
    files = result.get('files', [])
    return files[0]['id'] if files else None


def _list_receipt_candidates(drive, root_folder_id, date_str):
    """
    date_str(yyyy.mm.dd)로 시작하는 영수증 파일 목록 반환.

    폴더 구조를 세 가지 방식으로 탐색:
    1. root/"yyyy/mm" — 루트에 슬래시 포함 단일 폴더 (예: root/"2026/02")
    2. root/yyyy/mm — 중첩 폴더 (예: root/2026/02 또는 root/2026/2)
    3. root/yyyy/"yyyy/mm" — year 폴더 내 슬래시 포함 폴더 (예: root/2026/"2026/02")

    Returns list of {id, name, mimeType, webViewLink}.
    """
    parts = date_str.split('.')
    if len(parts) < 3:
        return []
    year, month = parts[0], parts[1]

    # 방식 1: "yyyy/mm" 이름의 단일 폴더
    month_id = _find_drive_subfolder(drive, root_folder_id, f"{year}/{month}")

    # 방식 2: 중첩 폴더 (year 폴더 → month 폴더)
    if not month_id:
        year_id = _find_drive_subfolder(drive, root_folder_id, year)
        if year_id:
            month_id = _find_drive_subfolder(drive, year_id, month)
            if not month_id:
                month_id = _find_drive_subfolder(drive, year_id, str(int(month)))
            # 방식 3: year 폴더 내에 "yyyy/mm" 이름의 폴더 (예: 2026/"2026/02")
            if not month_id:
                month_id = _find_drive_subfolder(drive, year_id, f"{year}/{month}")

    if not month_id:
        return []

    # 월 폴더 내 전체 파일을 받아 클라이언트 사이드에서 날짜 접두사 필터
    # (Drive API의 name contains 는 점(.)을 구분자로 처리하여 오매칭 발생)
    all_files = []
    page_token = None
    while True:
        kwargs = dict(
            q=f"'{month_id}' in parents and trashed=false",
            fields='nextPageToken, files(id, name, mimeType, webViewLink)',
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        )
        if page_token:
            kwargs['pageToken'] = page_token
        result = drive.files().list(**kwargs).execute()
        all_files.extend(result.get('files', []))
        page_token = result.get('nextPageToken')
        if not page_token:
            break

    return sorted(
        (f for f in all_files if f['name'].startswith(date_str)),
        key=lambda f: f['name'],
    )


def _export_drive_file_as_text(drive, file_id):
    """
    Google Drive 파일을 텍스트로 내보내기.

    text/plain → 실패 시 text/html(태그 제거) 순으로 시도.
    모두 실패하면 None 반환.
    """
    from googleapiclient.http import MediaIoBaseDownload

    for mime in ('text/plain', 'text/html'):
        buf = io.BytesIO()
        try:
            downloader = MediaIoBaseDownload(
                buf,
                drive.files().export_media(fileId=file_id, mimeType=mime),
            )
            done = False
            while not done:
                _, done = downloader.next_chunk()
        except Exception:
            continue

        text = buf.getvalue().decode('utf-8', errors='ignore')
        if mime == 'text/html':
            text = re.sub(r'<[^>]+>', '', text)
        return text

    return None


def _extract_amounts_from_drive_file(drive, file_id):
    """
    Google Drive 파일(Google Docs 등)을 텍스트로 내보내 포함된 정수 집합 반환.

    Returns set of int.
    """
    text = _export_drive_file_as_text(drive, file_id)
    if text is None:
        print(f"[WARNING] 파일 텍스트 내보내기 실패 (file_id={file_id}), 건너뜁니다.")
        return set()

    amounts = set()
    for m in re.finditer(r'([\d,]+)원', text):
        raw = m.group(1).replace(',', '')
        if raw:
            try:
                amounts.add(int(raw))
            except ValueError:
                pass
    return amounts


def _normalize_receipt_title(title):
    """영수증 파일 제목 정규화."""
    # 비기너 환급이 포함된 경우 → 비기너 환급
    if '비기너 환급' in title:
        return '비기너 환급'
    # "의 사본" 제거
    title = re.sub(r'의 사본\s*$', '', title).strip()
    # 이름 패턴 제거: 한글이름(트랙)님 또는 한글이름님
    title = re.sub(r'[가-힣]{2,5}(?:\s*\([A-Za-z]+\))?님\s*', '', title).strip()
    return title


def build_receipt_map(folder_url, transactions):
    """
    출금 거래 목록에 대해 영수증 파일 매칭 맵 생성.

    1차: 날짜(파일명 접두사 yyyy.mm.dd) 매칭
    2차: 파일 텍스트에 거래 금액(절댓값) 포함 여부 확인

    Returns:
        dict mapping (date_str, abs_amount_int) → (title, url)
        - title: 파일명에서 날짜 접두사와 확장자를 제거한 문자열
        - url:   Google Drive webViewLink
    """
    folder_id = _extract_drive_folder_id(folder_url)
    if not folder_id:
        raise ValueError(f"영수증 Drive 폴더 URL에서 ID를 파싱할 수 없습니다: {folder_url}")

    drive = _get_drive_service()
    receipt_map = {}

    # 동일 날짜·금액 출금이 2건 이상인 키는 어느 영수증인지 특정 불가 → 제외
    # date_str에 시간이 포함되어 있어도 날짜 부분(앞 10자)만 사용한다.
    tx_counts = Counter(
        (date_str[:10], int(abs(amount)))
        for date_str, amount, *_ in transactions
        if amount < 0
    )
    ambiguous = {key for key, cnt in tx_counts.items() if cnt > 1}

    withdrawal_dates = {date_str for date_str, *_ in tx_counts}

    for date_str in sorted(withdrawal_dates):
        candidates = _list_receipt_candidates(drive, folder_id, date_str)
        for f in candidates:
            amounts = _extract_amounts_from_drive_file(drive, f['id'])
            title = _normalize_receipt_title(f['name'][len(date_str):].strip())
            for amt in amounts:
                key = (date_str, amt)
                if key not in ambiguous and key not in receipt_map:
                    receipt_map[key] = (title, f['webViewLink'])
            # 이체 수수료 500원이 별도 기재된 경우: main + 500 키도 등록
            if 500 in amounts:
                for amt in amounts - {500}:
                    fee_key = (date_str, amt + 500)
                    if fee_key not in ambiguous and fee_key not in receipt_map:
                        receipt_map[fee_key] = (title, f['webViewLink'])

    return receipt_map


# ============================================================================
# Entry point
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="신한 거래내역 → 재학생 회비 관리 문서 자동 기입",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
자동 기입 항목: 날짜, 이름(거래내역 내용), 입/출, 잔액, 소계/합계 수식
수동 기입 항목: 내용(E열), 비고(G열)

예시:
  python fill_ledger.py
  python fill_ledger.py 신한_거래내역/신한_거래내역_2602.xlsx
  python fill_ledger.py --force
""",
    )
    parser.add_argument(
        "transaction_file",
        nargs="?",
        help="거래내역 파일 경로 (미지정 시 TRANSACTION_DRIVE_URL에서 다운로드)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="이미 데이터가 기입된 월도 강제로 덮어쓰기",
    )
    args = parser.parse_args()

    print("=" * 60)
    print("신한 거래내역 → 재학생 회비 관리 문서 자동 기입")
    print("=" * 60)

    tx_tmp_path = None

    try:
        # 거래내역 파일 결정
        tx_original_name = None
        if args.transaction_file:
            tx_file = args.transaction_file
        else:
            tx_drive_url = os.getenv('TRANSACTION_DRIVE_URL')
            if not tx_drive_url:
                print("[ERROR] TRANSACTION_DRIVE_URL 환경변수가 설정되지 않았습니다.")
                sys.exit(1)
            print("\n[INFO] 거래내역 Drive에서 다운로드 중...")
            try:
                tx_original_name, tx_tmp_path = download_transaction_from_drive(tx_drive_url)
                tx_file = tx_tmp_path
                print(f"[INFO] 다운로드 완료: {tx_original_name}")
            except (FileNotFoundError, ValueError, OSError, _HttpError) as e:
                print(f"[ERROR] 거래내역 Drive 다운로드 실패: {e}")
                sys.exit(1)

        print(f"\n[INFO] 거래내역 파일: {tx_original_name or tx_file}")

        # 연도/월 파싱: Drive에서 받은 경우 원본 파일명 기준
        try:
            year, month = get_year_month_from_filename(tx_original_name or tx_file)
        except ValueError as e:
            print(f"[ERROR] {e}")
            sys.exit(1)
        print(f"[INFO] 대상: {year}년 {month}월")

        # 관리 문서 URL
        management_sheet_url = os.getenv('MANAGEMENT_SHEET_URL')
        if not management_sheet_url:
            print("[ERROR] MANAGEMENT_SHEET_URL 환경변수가 설정되지 않았습니다.")
            sys.exit(1)

        print(f"[INFO] 관리 문서: {management_sheet_url}")
        spreadsheet_id = _extract_sheet_id(management_sheet_url)
        sheet_name = f"{year}년"

        # Sheets API 초기화 및 시트 확인
        try:
            sheets = _get_sheets_service()
            gid = _get_sheet_gid(sheets, spreadsheet_id, sheet_name)
        except (ValueError, _HttpError) as e:
            print(f"[ERROR] Google Sheets 접근 실패: {e}")
            sys.exit(1)

        print(f"[INFO] 시트 '{sheet_name}' 확인 완료")

        # 거래내역 파싱
        try:
            transactions = parse_transaction_file(tx_file)
        except Exception as e:
            print(f"[ERROR] 거래내역 파일이 손상되었거나 읽을 수 없습니다: {e}")
            sys.exit(1)
        print(f"[INFO] 파싱된 거래 건수: {len(transactions)}건")

        # 영수증 매칭
        receipt_map = {}
        receipt_drive_url = os.getenv('RECEIPT_DIR')
        if receipt_drive_url:
            print("\n[INFO] 영수증 Drive 폴더에서 매칭 중...")
            try:
                receipt_map = build_receipt_map(receipt_drive_url, transactions)
                print(f"[INFO] 영수증 매칭 완료: {len(receipt_map)}건")
            except Exception as e:
                print(f"[WARNING] 영수증 매칭 실패 (건너뜀): {e}")

        # 데이터 기입
        print()
        success = fill_month_api(
            sheets, spreadsheet_id, sheet_name, gid, month,
            transactions, force=args.force, receipt_map=receipt_map,
        )
        if success is False:
            sys.exit(1)
        if success is None:
            sys.exit(0)

        # 합계 수식 갱신
        update_total_formula_api(sheets, spreadsheet_id, sheet_name)

        print()
        print("=" * 60)
        print(f"[INFO] 기입 완료: {management_sheet_url}")
        print("[INFO] 아래 항목은 수동으로 기입해주세요:")
        if not receipt_drive_url:
            print("       - E열 (내용): 회비 / 서버비 / 회식비 등")
        else:
            print("       - E열 (내용): 영수증 미매칭 출금 항목")
        print("       - G열 (비고): 납부 월 등")
        print("=" * 60)

    finally:
        if tx_tmp_path and os.path.exists(tx_tmp_path):
            os.remove(tx_tmp_path)

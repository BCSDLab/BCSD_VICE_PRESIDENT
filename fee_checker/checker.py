#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BCSD 회비 납부 검증 및 미납 메시지 생성 프로그램
"""

import os
import sys
import re
import argparse
from contextlib import contextmanager
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from common.fee_notice import _render_fee_notice_message


# ============================================================================
# Constants
# ============================================================================

TEMPLATE_FILE = "templates/fee_notice.md"
OUTPUT_BASE_DIR = "output"

# 제외 키워드 (비고 열에서 검사)
EXCLUDE_KEYWORDS = ["졸업", "활동 중지", "휴학", "군 휴학", "트랙장", "교육장"]

# 엑셀 시트 설정
SHEETS_TO_PROCESS = ["2025", "2026"]
DATA_START_ROW = 5  # Row 5부터 데이터 시작

# 컬럼 매핑 (0-indexed)
COL_TRACK = 2  # C열: 트랙
COL_NAME = 3  # D열: 이름
COL_NOTES = 4  # E열: 비고
COL_MONTHS_START = 5  # F열: 1월 (F=5, G=6, ..., Q=16)

# 월별 컬럼 (F=1월, G=2월, ..., Q=12월)
MONTH_COLUMNS = {
    1: 5,  # F
    2: 6,  # G
    3: 7,  # H
    4: 8,  # I
    5: 9,  # J
    6: 10,  # K
    7: 11,  # L
    8: 12,  # M
    9: 13,  # N
    10: 14,  # O
    11: 15,  # P
    12: 16,  # Q
}

# 회비 금액
MONTHLY_FEE = 10_000   # ~2026년 2월: 월 10,000원
SEMESTER_FEE = 60_000  # 2026년 3월~: 학기당 60,000원

# 2026년 학기 정의: (시작 월, 종료 월)
# 실제 2학기는 9월~이듬해 2월이나, 2026 시트 내 구간(9~12월)만 처리
SEMESTERS_2026 = [
    (3, 8),   # 1학기
    (9, 12),  # 2학기 (2026 시트 내 구간)
]


# ============================================================================
# Helper Functions
# ============================================================================


def parse_sheet(ws, sheet_name):
    """
    시트에서 데이터 행 파싱

    Returns:
        list of dict: 각 행의 데이터
            {
                'track': str,
                'name': str,
                'notes': str,
                'months': {1: 'O'/'-'/None, 2: ..., 12: ...}
            }
    """
    rows = []
    row_idx = DATA_START_ROW

    while True:
        name_cell = ws.cell(row=row_idx, column=COL_NAME + 1)
        name = name_cell.value

        if name is None or str(name).strip() == "":
            break

        track_cell = ws.cell(row=row_idx, column=COL_TRACK + 1)
        notes_cell = ws.cell(row=row_idx, column=COL_NOTES + 1)

        track = track_cell.value if track_cell.value else ""
        notes = notes_cell.value if notes_cell.value else ""

        months = {}
        for month_num, col_idx in MONTH_COLUMNS.items():
            cell = ws.cell(row=row_idx, column=col_idx + 1)
            val = cell.value

            if val == "O":
                months[month_num] = "O"
            elif val == "-" or val == "−":
                months[month_num] = "-"
            else:
                months[month_num] = None

        rows.append(
            {
                "track": str(track).strip(),
                "name": str(name).strip(),
                "notes": str(notes).strip(),
                "months": months,
            }
        )

        row_idx += 1

    return rows


def should_exclude_row(row_data):
    """
    행 제외 여부 판단

    제외 조건:
    1. 비고에 제외 키워드 포함
    2. 모든 체크 대상 월이 "-" (면제)
    """
    notes = row_data["notes"]
    for keyword in EXCLUDE_KEYWORDS:
        if keyword in notes:
            return True

    months = row_data["months"]
    all_exempt = all(val == "-" for val in months.values())
    if all_exempt:
        return True

    return False


def calculate_unpaid_detail(row_data, sheet_name, current_month):
    """
    미납 상세 계산

    - 2025년 전체 / 2026년 1~2월: 월 10,000원
    - 2026년 3월~: 학기당 60,000원 (SEMESTERS_2026 참고)

    Args:
        row_data: parse_sheet() 반환 행 데이터
        sheet_name: "2025" 또는 "2026"
        current_month: 현재 월 (1~12)

    Returns:
        dict:
            monthly_amount   (int)       — 월별 미납 금액 합계
            unpaid_semesters (list[str]) — 미납 학기 식별자 목록 (예: ['26-1', '26-2'])
    """
    months = row_data["months"]

    if sheet_name == "2025":
        unpaid = sum(1 for m in range(1, 13) if months.get(m) is None)
        return {"monthly_amount": unpaid * MONTHLY_FEE, "unpaid_semesters": []}

    if sheet_name == "2026":
        check_until = max(0, current_month - 1)
        monthly_amount = 0
        unpaid_semesters = []

        # 1~2월: 월별 10,000원
        for m in range(1, min(2, check_until) + 1):
            if months.get(m) is None:
                monthly_amount += MONTHLY_FEE

        # 학기별: 체크 대상 구간 내 미납 월이 하나라도 있으면 학기비 전액 청구
        year_short = int(sheet_name) % 100
        for i, (sem_start, sem_end) in enumerate(SEMESTERS_2026, 1):
            checked = [m for m in range(sem_start, sem_end + 1) if m <= check_until]
            if checked and any(months.get(m) is None for m in checked):
                unpaid_semesters.append(f"{year_short}-{i}")

        return {"monthly_amount": monthly_amount, "unpaid_semesters": unpaid_semesters}

    # 기타 시트
    unpaid = sum(1 for m in range(1, 13) if months.get(m) is None)
    return {"monthly_amount": unpaid * MONTHLY_FEE, "unpaid_semesters": []}


def _format_unpaid_detail(name, data, date_year, date_month, date_day):
    """
    미납 내역 문구 생성 (3가지 케이스)

    Case 1 — 월별 회비만 미납:
        "{이름} 님의 회비가 {금액}원 미납되었습니다."
    Case 2 — 학기 회비만 미납:
        "{이름} 님의 26-1학기 회비가 미납되었습니다."
    Case 3 — 월별 + 학기 혼재:
        "{이름} 님의 회비가 총 {금액}원 미납되었습니다.
          - 이전 회비: {월별금액}원
          - 26-1학기 회비: 60,000원"
    """
    monthly_amount = data["monthly_amount"]
    unpaid_semesters = data["unpaid_semesters"]
    prefix = f"확인 결과, {date_year}년 {date_month}월 {date_day}일 기준으로 {name} 님의"

    has_monthly = monthly_amount > 0
    has_semester = bool(unpaid_semesters)

    if has_monthly and not has_semester:
        return f"{prefix} 회비가 {monthly_amount:,}원 미납되었습니다."

    if not has_monthly and has_semester:
        sem_str = ", ".join(f"{s}학기" for s in unpaid_semesters)
        return f"{prefix} {sem_str} 회비가 미납되었습니다."

    if has_monthly and has_semester:
        total = data["unpaid_amount"]
        lines = [f"{prefix} 회비가 총 {total:,}원 미납되었습니다."]
        lines.append(f"  - 이전 회비: {monthly_amount:,}원")
        for s in unpaid_semesters:
            lines.append(f"  - {s}학기 회비: {SEMESTER_FEE:,}원")
        return "\n".join(lines) + "\n"

    return f"{prefix} 회비가 미납되었습니다."


def _validate_identifier(value):
    """SQL 식별자 검증: 영문자·숫자·언더스코어만 허용 (SQL 인젝션 방지)"""
    if not re.fullmatch(r'[A-Za-z_][A-Za-z0-9_]*', value):
        raise ValueError(f"유효하지 않은 SQL 식별자: '{value}'")
    return value


def _normalize_name(name):
    """이름 정규화: 앞뒤 공백 제거"""
    if name is None:
        return ""
    return str(name).strip()


def _normalize_track(track):
    """트랙명 정규화: 영문자만 추출 후 소문자 변환 (예: 'FrontEnd' → 'frontend')"""
    if track is None:
        return ""
    return re.sub(r'[^a-zA-Z]', '', track).lower()


@contextmanager
def _ssh_tunnel(ssh_host, ssh_port, ssh_user, remote_host, remote_port, ssh_key_path=None, ssh_password=None):
    """paramiko 기반 SSH 포트 포워딩 터널"""
    import select
    import threading
    import socketserver
    import paramiko

    client = paramiko.SSHClient()
    client.load_system_host_keys()
    client.set_missing_host_key_policy(paramiko.RejectPolicy())

    connect_kwargs: dict = {"port": ssh_port, "username": ssh_user}
    if ssh_key_path:
        connect_kwargs["key_filename"] = os.path.expanduser(ssh_key_path)
    else:
        connect_kwargs["password"] = ssh_password

    client.connect(ssh_host, **connect_kwargs)
    transport = client.get_transport()
    if transport is None:
        client.close()
        raise RuntimeError("SSH 연결 후 transport를 가져올 수 없습니다.")

    class _ForwardHandler(socketserver.BaseRequestHandler):
        def handle(self):
            chan = transport.open_channel(
                "direct-tcpip",
                (remote_host, remote_port),
                self.request.getpeername(),
            )
            if chan is None:
                return
            while True:
                readable = select.select([self.request, chan], [], [], 5)[0]
                if self.request in readable:
                    data = self.request.recv(1024)
                    if not data:
                        break
                    chan.send(data)
                if chan in readable:
                    data = chan.recv(1024)
                    if not data:
                        break
                    self.request.send(data)
            chan.close()

    server = None
    try:
        server = socketserver.ThreadingTCPServer(("127.0.0.1", 0), _ForwardHandler)
        local_port = server.server_address[1]
        threading.Thread(target=server.serve_forever, daemon=True).start()
        yield local_port
    finally:
        if server is not None:
            server.shutdown()
            server.server_close()
        client.close()


@contextmanager
def _db_connection():
    """SSH 터널 경유 MySQL 연결 컨텍스트 매니저 (readonly)"""
    try:
        import pymysql
    except ImportError as err:
        raise ImportError("필요한 패키지: pip install pymysql paramiko") from err

    ssh_host = os.getenv("SSH_HOST", "")
    try:
        ssh_port = int(os.getenv("SSH_PORT", "22"))
    except ValueError:
        raise ValueError("SSH_PORT 환경 변수가 유효한 정수가 아닙니다.")
    ssh_user = os.getenv("SSH_USER", "")
    ssh_key_path = os.getenv("SSH_KEY_PATH")
    ssh_password = os.getenv("SSH_PASSWORD", "")

    db_host = os.getenv("DB_HOST", "127.0.0.1")
    try:
        db_port = int(os.getenv("DB_PORT", "3306"))
    except ValueError:
        raise ValueError("DB_PORT 환경 변수가 유효한 정수가 아닙니다.")
    db_name = os.getenv("DB_NAME", "")
    db_user = os.getenv("DB_USER", "")
    db_password = os.getenv("DB_PASSWORD", "")

    if not ssh_host:
        raise ValueError("SSH_HOST 환경 변수가 설정되지 않았습니다.")
    if not ssh_user:
        raise ValueError("SSH_USER 환경 변수가 설정되지 않았습니다.")

    with _ssh_tunnel(ssh_host, ssh_port, ssh_user, db_host, db_port, ssh_key_path, ssh_password) as local_port:
        conn = pymysql.connect(
            host="127.0.0.1",
            port=local_port,
            user=db_user,
            password=db_password,
            database=db_name,
            charset="utf8mb4",
            cursorclass=pymysql.cursors.DictCursor,
        )
        try:
            yield conn
        finally:
            conn.close()


def fetch_slack_id_map():
    """DB에서 (이름, 트랙명) → Slack ID 매핑 조회 (readonly SELECT)"""
    table          = _validate_identifier(os.getenv("DB_TABLE", ""))
    col_name       = _validate_identifier(os.getenv("DB_COL_NAME", "name"))
    col_slack_id   = _validate_identifier(os.getenv("DB_COL_SLACK_ID", "slack_id"))
    col_track_id   = _validate_identifier(os.getenv("DB_COL_TRACK_ID", "track_id"))
    col_is_deleted       = _validate_identifier(os.getenv("DB_COL_IS_DELETED", "is_deleted"))
    track_table          = _validate_identifier(os.getenv("DB_TRACK_TABLE", ""))
    track_col_id         = _validate_identifier(os.getenv("DB_TRACK_COL_ID", "id"))
    track_col_name       = _validate_identifier(os.getenv("DB_TRACK_COL_NAME", "name"))
    track_col_is_deleted = _validate_identifier(
        os.getenv("DB_TRACK_COL_IS_DELETED", os.getenv("DB_COL_IS_DELETED", "is_deleted"))
    )

    with _db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT m.`{col_name}`, t.`{track_col_name}` AS track_name, m.`{col_slack_id}`"
                f" FROM `{table}` m"
                f" JOIN `{track_table}` t ON m.`{col_track_id}` = t.`{track_col_id}`"
                f" WHERE m.`{col_slack_id}` IS NOT NULL AND m.`{col_slack_id}` != ''"
                f" AND m.`{col_is_deleted}` = 0 AND t.`{track_col_is_deleted}` = 0"
            )
            return {
                (_normalize_name(row[col_name]), _normalize_track(row["track_name"])): row[col_slack_id]
                for row in cur.fetchall()
            }


def aggregate_unpaid_fees(wb, current_month, excluded_tracks=None, excluded_persons=None):
    """
    2025/2026 시트 데이터 통합 및 미납 금액 계산

    Returns:
        dict: {(name, track): {'name': str, 'track': str, 'unpaid_amount': int,
                               'monthly_amount': int, 'unpaid_semesters': list[str]}}
    """
    if excluded_tracks is None:
        excluded_tracks = set()
    else:
        excluded_tracks = set(excluded_tracks)
    if excluded_persons is None:
        excluded_persons = set()
    else:
        excluded_persons = set(excluded_persons)

    aggregated = {}
    excluded_keys = set()
    exempt_names_from_2025 = set()

    for sheet_name in SHEETS_TO_PROCESS:
        if sheet_name not in wb.sheetnames:
            print(f"[WARNING] 시트 '{sheet_name}' 없음, 건너뜀")
            continue

        ws = wb[sheet_name]
        rows = parse_sheet(ws, sheet_name)

        print(f"\n[INFO] 시트 '{sheet_name}': {len(rows)}개 행 파싱됨")

        included_count = 0
        excluded_count = 0

        if sheet_name == "2025":
            for row_data in rows:
                notes = row_data["notes"]
                for keyword in EXCLUDE_KEYWORDS:
                    if keyword in notes:
                        exempt_names_from_2025.add(row_data["name"])
                        break

        for row_data in rows:
            name = row_data["name"]
            track = row_data["track"]
            key = (name, track)

            if name in exempt_names_from_2025:
                excluded_keys.add(key)
                excluded_count += 1
                continue

            if _normalize_track(track) in excluded_tracks:
                excluded_keys.add(key)
                excluded_count += 1
                continue

            if (name, _normalize_track(track)) in excluded_persons:
                excluded_keys.add(key)
                excluded_count += 1
                continue

            if should_exclude_row(row_data):
                excluded_keys.add(key)
                excluded_count += 1
                continue

            included_count += 1

            detail = calculate_unpaid_detail(row_data, sheet_name, current_month)
            unpaid_amount = detail["monthly_amount"] + len(detail["unpaid_semesters"]) * SEMESTER_FEE

            if key not in aggregated:
                aggregated[key] = {
                    "name": name,
                    "track": track,
                    "unpaid_amount": 0,
                    "monthly_amount": 0,
                    "unpaid_semesters": [],
                }

            aggregated[key]["unpaid_amount"] += unpaid_amount
            aggregated[key]["monthly_amount"] += detail["monthly_amount"]
            aggregated[key]["unpaid_semesters"].extend(detail["unpaid_semesters"])

        print(f"[INFO] 시트 '{sheet_name}': 포함 {included_count}명, 제외 {excluded_count}명")

    for excluded_key in excluded_keys:
        aggregated.pop(excluded_key, None)

    filtered = {k: v for k, v in aggregated.items() if v["unpaid_amount"] > 0}
    return filtered


# ============================================================================
# Main Functions
# ============================================================================


def get_output_directory():
    """현재 연-월 기준으로 output 디렉토리 생성 및 반환 (예: output/2026-02/)"""
    now = datetime.now()
    output_dir = os.path.join(OUTPUT_BASE_DIR, f"{now.year:04d}-{now.month:02d}")
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


def _previous_month_last_day():
    """전월 말일 datetime 객체 반환"""
    return datetime.now().replace(day=1) - timedelta(days=1)


def get_previous_month_end_date():
    """전월 말일을 한국어 형식으로 반환 (예: "2026년 1월 31일")"""
    last_day = _previous_month_last_day()
    return f"{last_day.year}년 {last_day.month}월 {last_day.day}일"


def format_amount(amount):
    """금액을 천 단위 콤마 형식으로 변환 (예: 120000 → "120,000")"""
    return f"{amount:,}"


def generate_unique_filename(name, track, used_filenames):
    """중복 이름 처리를 위한 고유 파일명 생성"""
    track_filename = f"{name}_{track}.txt"

    if track_filename not in used_filenames:
        used_filenames.add(track_filename)
        return track_filename

    counter = 1
    while True:
        numbered_filename = f"{name}_{track}_{counter}.txt"
        if numbered_filename not in used_filenames:
            used_filenames.add(numbered_filename)
            return numbered_filename
        counter += 1


def generate_message_files(unpaid_data, output_dir, template_path):
    """
    미납 회원별 메시지 파일 생성

    Returns:
        tuple: (생성된 파일 수, 총 미납 금액)
    """
    for existing_file in Path(output_dir).glob("*.txt"):
        existing_file.unlink()

    with open(template_path, "r", encoding="utf-8") as f:
        template_content = f.read()

    last_day = _previous_month_last_day()

    used_filenames = set()
    files_generated = 0
    total_unpaid_amount = 0

    sender_name = os.getenv("SENDER_NAME", "")
    sender_phone = os.getenv("SENDER_PHONE", "")
    fee_sheet_url = os.getenv("FEE_SHEET_URL", "")
    if not fee_sheet_url:
        print("[WARNING] FEE_SHEET_URL 환경 변수가 설정되지 않았습니다. 납부문서 링크가 비어 있습니다.")

    for (name, track), data in unpaid_data.items():
        unpaid_detail = _format_unpaid_detail(name, data, last_day.year, last_day.month, last_day.day)

        message = _render_fee_notice_message(
            template_content=template_content,
            sender_name=sender_name,
            sender_phone=sender_phone,
            mention=f"@{sender_name}" if sender_name else "{멘션}",
            unpaid_detail=unpaid_detail,
            fee_sheet_url=fee_sheet_url,
        )

        filename = generate_unique_filename(name, track, used_filenames)
        filepath = os.path.join(output_dir, filename)

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(message)

        files_generated += 1
        total_unpaid_amount += data["unpaid_amount"]

    return files_generated, total_unpaid_amount


def send_slack_dms(unpaid_data, template_path):
    """
    미납 회원에게 Slack DM 발송

    Returns:
        tuple: (발송 성공 수, 실패/미매칭 수)
    """
    try:
        from slack_sdk import WebClient
        from slack_sdk.errors import SlackApiError
    except ImportError as err:
        raise ImportError("slack_sdk 패키지가 필요합니다: pip install slack-sdk") from err

    token = os.getenv("SLACK_BOT_TOKEN")
    if not token:
        raise ValueError("SLACK_BOT_TOKEN 환경 변수가 설정되지 않았습니다.")

    sender_id = os.getenv("SLACK_SENDER_ID", "")
    if not sender_id:
        raise ValueError("SLACK_SENDER_ID 환경 변수가 설정되지 않았습니다.")

    sender_name = os.getenv("SENDER_NAME", "")
    sender_phone = os.getenv("SENDER_PHONE", "")
    fee_sheet_url = os.getenv("FEE_SHEET_URL", "")
    if not sender_name:
        raise ValueError("SENDER_NAME 환경 변수가 설정되지 않았습니다.")
    if not sender_phone:
        raise ValueError("SENDER_PHONE 환경 변수가 설정되지 않았습니다.")
    if not fee_sheet_url:
        raise ValueError("FEE_SHEET_URL 환경 변수가 설정되지 않았습니다.")

    client = WebClient(token=token)

    with open(template_path, "r", encoding="utf-8") as f:
        template_content = f.read()

    last_day = _previous_month_last_day()

    print("[INFO] DB에서 Slack ID 조회 중...")
    name_to_user_id = fetch_slack_id_map()
    print(f"[INFO] 조회된 멤버 수: {len(name_to_user_id)}명")

    sent = 0
    failed = 0

    for (name, track), data in unpaid_data.items():
        user_id = name_to_user_id.get((_normalize_name(name), _normalize_track(track)))
        if not user_id:
            print(f"[WARNING] Slack 유저를 찾을 수 없음: {name} ({track})")
            failed += 1
            continue

        unpaid_detail = _format_unpaid_detail(name, data, last_day.year, last_day.month, last_day.day)
        message = _render_fee_notice_message(
            template_content=template_content,
            sender_name=sender_name,
            sender_phone=sender_phone,
            mention=f"<@{sender_id}>",
            unpaid_detail=unpaid_detail,
            fee_sheet_url=fee_sheet_url,
        )

        try:
            dm_resp = client.conversations_open(users=[user_id])
            channel_id = dm_resp["channel"]["id"]
            client.chat_postMessage(channel=channel_id, text=message)
            print(f"[INFO] DM 발송 완료: {name} ({track})")
            sent += 1
        except SlackApiError as e:
            print(f"[ERROR] DM 발송 실패: {name} ({track}) - {e.response['error']}")
            failed += 1
        except Exception as e:
            print(f"[ERROR] DM 발송 중 예외 발생: {name} ({track}) - {e}")
            failed += 1

    return sent, failed


def parse_excluded_tracks(exclude_args):
    """CLI 인자에서 제외할 트랙 파싱 (영소문자 정규화)"""
    excluded_tracks = set()
    if exclude_args:
        for arg in exclude_args:
            tracks = [_normalize_track(t.strip()) for t in arg.split(",")]
            excluded_tracks.update(tracks)
    return excluded_tracks


def parse_excluded_persons(exclude_args):
    """CLI 인자에서 제외할 개인 파싱 (이름_트랙 형식, 트랙 영소문자 정규화)"""
    excluded_persons = set()
    if exclude_args:
        for arg in exclude_args:
            for entry in arg.split(","):
                entry = entry.strip()
                if "_" not in entry:
                    print(f"[WARNING] 개인 제외 형식 오류 (이름_트랙 필요): '{entry}'")
                    continue
                name, track = entry.split("_", 1)
                excluded_persons.add((name.strip(), _normalize_track(track)))
    return excluded_persons


def main():
    parser = argparse.ArgumentParser(description="BCSD 회비 납부 검증 및 미납 메시지 생성 프로그램")
    parser.add_argument(
        "-e",
        "--exclude-track",
        action="append",
        dest="exclude_tracks",
        help="제외할 트랙명 (반복 사용 가능, 쉼표로 구분 가능)",
    )
    parser.add_argument(
        "-p",
        "--exclude-person",
        action="append",
        dest="exclude_persons",
        help="제외할 개인 — 이름_트랙 형식 (반복 사용 가능, 쉼표로 구분 가능)",
    )
    parser.add_argument(
        "--send-dm",
        action="store_true",
        help=(
            "미납 회원에게 Slack DM 발송. "
            "필요한 환경 변수: SLACK_BOT_TOKEN, SLACK_SENDER_ID, SENDER_NAME, SENDER_PHONE, "
            "SSH_HOST, SSH_USER, SSH_KEY_PATH (또는 SSH_PASSWORD), "
            "DB_TABLE, DB_TRACK_TABLE (및 기타 DB_* 변수)"
        ),
    )
    args = parser.parse_args()

    excluded_tracks = parse_excluded_tracks(args.exclude_tracks)
    excluded_persons = parse_excluded_persons(args.exclude_persons)

    print("=" * 70)
    print("BCSD 회비 납부 검증 및 미납 메시지 생성 프로그램")
    print("=" * 70)

    output_dir = get_output_directory()
    print(f"\n[INFO] 출력 디렉토리: {output_dir}")

    if excluded_tracks:
        print(f"[INFO] 제외할 트랙: {', '.join(sorted(excluded_tracks))}")
    if excluded_persons:
        print(f"[INFO] 제외할 개인: {', '.join(f'{n}_{t}' for n, t in sorted(excluded_persons))}")

    fee_sheet_url = os.getenv("FEE_SHEET_URL")
    if not fee_sheet_url:
        print("[ERROR] FEE_SHEET_URL 환경변수가 설정되지 않았습니다.", file=sys.stderr)
        sys.exit(1)

    if not os.path.exists(TEMPLATE_FILE):
        print(f"[ERROR] 템플릿 파일을 찾을 수 없습니다 ({TEMPLATE_FILE})")
        sys.exit(1)

    from ledger_filler.filler import download_sheet_as_xlsx
    print(f"[INFO] 납부 문서 다운로드 중... ({fee_sheet_url})")
    _, tmp_path = download_sheet_as_xlsx(fee_sheet_url)
    print(f"[INFO] 다운로드 완료 → {tmp_path}")
    print(f"[INFO] 선택된 템플릿 파일: {TEMPLATE_FILE}")

    print("\n" + "=" * 70)
    print("엑셀 파싱 및 미납 계산")
    print("=" * 70)

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=False)
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
    current_month = datetime.now().month

    print(f"[INFO] 현재 월: {current_month}")
    print(f"[INFO] 처리 대상 시트: {SHEETS_TO_PROCESS}")

    unpaid_data = aggregate_unpaid_fees(wb, current_month, excluded_tracks, excluded_persons)

    print("\n" + "=" * 70)
    print("미납 데이터 요약")
    print("=" * 70)
    print(f"[INFO] 총 미납 대상자: {len(unpaid_data)}명")

    if unpaid_data:
        print("\n미납 데이터 샘플 (최대 5명):")
        for idx, ((name, track), data) in enumerate(unpaid_data.items()):
            if idx >= 5:
                break
            print(f"  - {name} ({track}): {data['unpaid_amount']:,}원")

    print("\n" + "=" * 70)
    print("메시지 생성 및 파일 출력")
    print("=" * 70)

    files_generated, total_unpaid_amount = generate_message_files(
        unpaid_data, output_dir, TEMPLATE_FILE
    )

    print(f"\n[INFO] 생성된 파일 수: {files_generated}개")
    print(f"[INFO] 총 미납 금액: {format_amount(total_unpaid_amount)}원")

    if args.send_dm and unpaid_data:
        print("\n" + "=" * 70)
        print("Slack DM 발송")
        print("=" * 70)
        dm_sent, dm_failed = send_slack_dms(unpaid_data, TEMPLATE_FILE)
        print(f"[INFO] DM 발송 완료: {dm_sent}명, 실패/미매칭: {dm_failed}명")

    print("\n" + "=" * 70)
    print("처리 완료 요약")
    print("=" * 70)
    print(f"[INFO] 처리된 시트: {', '.join(SHEETS_TO_PROCESS)}")
    print(f"[INFO] 미납 대상자: {len(unpaid_data)}명")
    print(f"[INFO] 생성된 메시지 파일: {files_generated}개")
    print(f"[INFO] 총 미납 금액: {format_amount(total_unpaid_amount)}원")
    print(f"[INFO] 출력 디렉토리: {output_dir}")
    print("=" * 70)
